import streamlit as st
import pandas as pd
import io
import os
import time
import requests
import concurrent.futures
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import Footballtravel   
import Olka 
import Fantravel 
import Fodboldrejseguiden  

def send_discord_notification(selected_clubs, execution_time):
    try:
        webhook_url = st.secrets["DISCORD_WEBHOOK_URL"]
        
        payload = {
            "embeds": [{
                "title": "Scraper Eksekveret",
                "color": 3066993,
                "fields": [
                    {"name": "Klubber valgt", "value": ", ".join(selected_clubs), "inline": False},
                    {"name": "Tid brugt", "value": execution_time, "inline": True},
                    {"name": "Status", "value": "Excel genereret", "inline": True}
                ],
                "timestamp": datetime.now().isoformat()
            }]
        }
        
        requests.post(webhook_url, json=payload)
    except Exception as e:
        print(f"Discord fejl: {e}")

st.set_page_config(page_title="Football Scraper Pro", layout="wide")

st.markdown("""
<style>
    div[data-testid="stButton"] > button[kind="primary"] {
        background-color: #28a745 !important;
        border-color: #28a745 !important; 
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

def get_club_names():
    if not os.path.exists("club_names.xlsx"):
        st.error("Mangler 'club_names.xlsx'")
        return []
    try:
        return pd.read_excel("club_names.xlsx", sheet_name="EN", usecols="A", header=None)[0].dropna().astype(str).str.strip().drop_duplicates().tolist()
    except: 
        return []

def main():
    st.title("Prissammenligning: Billet + Hotel")
    
    excel_clubs = get_club_names()
    if "selected_clubs" not in st.session_state: 
        st.session_state.selected_clubs = set()

    cols = st.columns(4)
    for i, club in enumerate(excel_clubs):
        if cols[i%4].button(club, key=club, type="primary" if club in st.session_state.selected_clubs else "secondary", use_container_width=True):
            if club in st.session_state.selected_clubs: 
                st.session_state.selected_clubs.remove(club)
            else: 
                st.session_state.selected_clubs.add(club)
            st.rerun()

    selected = list(st.session_state.selected_clubs)

    if selected:
        st.divider()
        if st.button("Søg efter priser", type="primary"):
            
            start_time = time.time()
            status = st.status("Indsamler data...", expanded=True)
            
            results = {
                "Footballtravel": pd.DataFrame(),
                "Olka": pd.DataFrame(),
                "Fantravel": pd.DataFrame(),
                "Resterende": pd.DataFrame()
            }

            try:
                results["Olka"] = Olka.get_prices(selected)
                status.write(f"Data hentet fra Olka: {len(results['Olka'])} tilbud fundet")
            except Exception as e:
                st.error(f"Fejl i Olka: {e}")

            with concurrent.futures.ThreadPoolExecutor() as executor:
                futures = {
                    executor.submit(Footballtravel.get_prices, selected): "Footballtravel",
                    executor.submit(Fantravel.get_prices, selected): "Fantravel",
                    executor.submit(Fodboldrejseguiden.get_prices, selected): "Resterende"
                }
                
                for future in concurrent.futures.as_completed(futures):
                    provider = futures[future]
                    try:
                        data = future.result()
                        if provider == "Footballtravel" and not data.empty:
                            data['Provider'] = "Footballtravel.dk"
                        results[provider] = data
                        status.write(f"Data hentet fra {provider}: {len(data)} tilbud fundet")
                    except Exception as e:
                        st.error(f"Fejl i {provider}: {e}")

            df1 = results["Footballtravel"]
            df2 = results["Olka"]
            df3 = results["Fantravel"]
            df5 = results["Resterende"]

            end_time = time.time()
            elapsed = int(end_time - start_time)
            mins, secs = divmod(elapsed, 60)
            time_str = f"{mins}m {secs}s"
            
            send_discord_notification(selected, time_str)
            status.update(label=f"Færdig (Tid: {time_str})", state="complete", expanded=False)
            
            frames = [df1, df2, df3, df5]
            if all(df.empty for df in frames):
                st.warning("Ingen priser fundet.")
                st.stop()
            
            full_df = pd.concat(frames, ignore_index=True)

            full_df['Provider'] = full_df['Provider'].fillna("Ukendt").astype(str)
            full_df = full_df[full_df['Provider'].str.strip() != ""]
            full_df['SortDate'] = pd.to_datetime(full_df['SortDate'], errors='coerce')
            full_df = full_df.dropna(subset=['SortDate'])

            cutoff = datetime.now() + timedelta(hours=24)
            full_df = full_df[full_df['SortDate'] > cutoff]
            if full_df.empty:
                st.warning("Ingen relevante kampe fundet.")
                st.stop()

            full_df = full_df.sort_values(by=['Club', 'SortDate'])
            full_df['club_change'] = full_df['Club'] != full_df['Club'].shift()
            full_df['date_diff'] = full_df['SortDate'].diff().dt.days.abs()
            full_df['big_gap'] = full_df['date_diff'] > 2 
            full_df['Match_Group_ID'] = (full_df['club_change'] | full_df['big_gap']).cumsum()
            
            all_providers = sorted(full_df['Provider'].unique())
            if "Footballtravel.dk" in all_providers:
                all_providers.remove("Footballtravel.dk")
                all_providers.insert(0, "Footballtravel.dk")

            matches_grouped = full_df.groupby('Match_Group_ID').agg({
                'Club': 'first',
                'Match': lambda x: max(x, key=len),
                'SortDate': 'first'
            }).reset_index()

            match_data_list = []
            
            for _, match_row in matches_grouped.iterrows():
                group_id = match_row['Match_Group_ID']
                match_name = match_row['Match']
                date_str = match_row['SortDate'].strftime('%d/%m')
                display_name = f"{match_name} ({date_str})"
                
                prices_in_group = full_df[full_df['Match_Group_ID'] == group_id]
                
                provider_data = {}
                
                for _, p_row in prices_in_group.iterrows():
                    prov = p_row['Provider']
                    price_val = p_row['Price']
                    nights_val = p_row['Nights']
                    
                    if prov not in provider_data or price_val < provider_data[prov]['price']:
                        provider_data[prov] = {
                            'price': price_val,
                            'nights': nights_val
                        }
                
                final_prices = [d['price'] for d in provider_data.values() if d['price'] > 0]
                
                min_price = min(final_prices) if final_prices else None
                max_price = max(final_prices) if final_prices else None

                match_data_list.append({
                    'display': display_name,
                    'data': provider_data,
                    'min_price': min_price, 
                    'max_price': max_price,
                    'club': match_row['Club']
                })

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                writer.book.create_sheet('Prices')
                ws = writer.book['Prices']
                
                ws.sheet_view.showGridLines = False

                header_font = Font(bold=True)
                header_alignment = Alignment(textRotation=45, vertical='bottom', horizontal='center')
                
                thin_side = Side(style='thin')
                medium_side = Side(style='medium') 
                
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                thick_left_border = Border(left=medium_side, right=thin_side, top=thin_side, bottom=thin_side)

                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                col_idx = 2
                prev_club = None 

                for match in match_data_list:
                    current_club = match['club']
                    use_thick_border = (prev_club is not None) and (current_club != prev_club)
                    current_border = thick_left_border if use_thick_border else thin_border
                    
                    prev_club = current_club

                    cell_match = ws.cell(row=1, column=col_idx, value=match['display'])
                    cell_match.font = header_font
                    cell_match.alignment = header_alignment
                    cell_match.border = current_border 
                    
                    cell_nights = ws.cell(row=1, column=col_idx+1, value="Nætter")
                    cell_nights.font = header_font
                    cell_nights.alignment = header_alignment
                    cell_nights.border = thin_border 
                    
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    ws.column_dimensions[get_column_letter(col_idx+1)].width = 8
                    
                    col_idx += 2

                row_idx = 2
                
                for provider in all_providers:
                    cell_prov = ws.cell(row=row_idx, column=1, value=provider)
                    cell_prov.font = Font(bold=True)
                    cell_prov.border = Border(top=medium_side, bottom=medium_side, left=medium_side, right=medium_side)
                    
                    col_idx = 2
                    prev_club = None 

                    for match in match_data_list:
                        current_club = match['club']
                        use_thick_border = (prev_club is not None) and (current_club != prev_club)
                        current_border = thick_left_border if use_thick_border else thin_border
                        prev_club = current_club

                        p_data = match['data'].get(provider, {'price': 0, 'nights': 0})
                        price = p_data['price']
                        nights = p_data['nights']
                        
                        cell_p = ws.cell(row=row_idx, column=col_idx, value=price if price > 0 else "")
                        cell_p.border = current_border 
                        
                        if price > 0:
                            if price == match['min_price']:
                                cell_p.fill = green_fill
                            elif price == match['max_price']:
                                cell_p.fill = red_fill
                        
                        cell_n = ws.cell(row=row_idx, column=col_idx+1, value=nights if nights > 0 else "")
                        cell_n.border = thin_border 
                        
                        col_idx += 2
                    
                    row_idx += 1
                
                row_idx += 2

                sect_header = ws.cell(row=row_idx, column=1, value="Sammenligning med Footballtravel.dk. Grøn = dyrere, Rød = billigere")
                sect_header.font = Font(bold=True, size=11)
                sect_header.border = Border(bottom=medium_side)
                row_idx += 1

                for provider in all_providers:
                    cell_prov = ws.cell(row=row_idx, column=1, value=provider)
                    cell_prov.font = Font(bold=True)
                    cell_prov.border = Border(top=medium_side, bottom=medium_side, left=medium_side, right=medium_side)
                    
                    col_idx = 2
                    prev_club = None

                    for match in match_data_list:
                        current_club = match['club']
                        use_thick_border = (prev_club is not None) and (current_club != prev_club)
                        current_border = thick_left_border if use_thick_border else thin_border
                        prev_club = current_club

                        p_data = match['data'].get(provider, {'price': 0, 'nights': 0})
                        price = p_data['price']
                        nights = p_data['nights']

                        ft_data = match['data'].get("Footballtravel.dk", {'price': 0, 'nights': 0})
                        ft_nights = ft_data['nights']
                        ft_price = ft_data['price']

                        price_diff_val = ""
                        if price > 0 and ft_price > 0:
                            price_diff_val = price - ft_price                       

                        nights_diff_val = ""
                        if price > 0 and ft_nights > 0:
                            nights_diff_val = nights - ft_nights
                        
                        cell_p_diff = ws.cell(row=row_idx, column=col_idx, value=price_diff_val)
                        cell_p_diff.border = current_border

                        if isinstance(price_diff_val, (int, float)):
                            if price_diff_val > 0:
                                cell_p_diff.fill = green_fill
                            elif price_diff_val < 0:
                                cell_p_diff.fill = red_fill
                        
                        cell_n_diff = ws.cell(row=row_idx, column=col_idx+1, value=nights_diff_val)
                        cell_n_diff.border = thin_border

                        col_idx += 2
                    
                    row_idx += 1

                ws.column_dimensions['A'].width = 25
                ws.freeze_panes = "B2"

            timestamp = datetime.now().strftime("%H-%M")
            st.download_button(
                "📥 Download Excel", 
                output.getvalue(), 
                f"prices_matrix_{timestamp}.xlsx", 
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            preview_df = pd.DataFrame(index=all_providers)
            for m in match_data_list:
                col_name = m['display']
                prices = []
                for p in all_providers:
                    val = m['data'].get(p, {}).get('price', 0)
                    prices.append(val if val > 0 else 0)
                preview_df[col_name] = prices
            
            st.write("Preview af data:")
            st.dataframe(preview_df, use_container_width=True)

if __name__ == "__main__":
    main()