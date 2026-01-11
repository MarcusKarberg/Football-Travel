import streamlit as st
import pandas as pd
import io
import os
import time
import subprocess
import sys

# Tjekker for Playwright
try:
    from playwright.sync_api import sync_playwright
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "playwright"])
    subprocess.run(["playwright", "install", "chromium"])

from datetime import datetime, timedelta
# Tilføjet 'Alignment' til imports for at kunne rotere tekst
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- IMPORTER VORES MODULER ---
# Sørg for at filerne (Footballtravel.py, Olka.py, osv.) ligger i samme mappe
import Footballtravel   
import Olka 
import Fantravel 
import Fodboldrejseguiden  

st.set_page_config(page_title="Football Scraper Pro", layout="wide")

# Styling
st.markdown("""
<style>
    div[data-testid="stButton"] > button[kind="primary"] {
        background-color: #28a745 !important;
        border-color: #28a745 !important; 
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)

def get_club_names():
    if not os.path.exists("club_names.xlsx"):
        st.error("❌ Mangler 'club_names.xlsx'")
        return []
    try:
        return pd.read_excel("club_names.xlsx", sheet_name="EN", usecols="A", header=None)[0].dropna().astype(str).str.strip().tolist()
    except: return []

def main():
    st.title("⚽ Prissammenligning: Billet + Hotel")
    
    excel_clubs = get_club_names()
    if "selected_clubs" not in st.session_state: st.session_state.selected_clubs = set()

    # Vælg klubber
    cols = st.columns(4)
    for i, club in enumerate(excel_clubs):
        if cols[i%4].button(club, key=club, type="primary" if club in st.session_state.selected_clubs else "secondary", use_container_width=True):
            if club in st.session_state.selected_clubs: st.session_state.selected_clubs.remove(club)
            else: st.session_state.selected_clubs.add(club)
            st.rerun()

    selected = list(st.session_state.selected_clubs)

    if selected:
        st.divider()
        if st.button("🔎 Søg efter priser", type="primary"):
            
            # --- START TIMER ---
            start_time = time.time()
            
            # Vægtning af tid (til progress bar)
            P_FT = 60
            P_OLKA = 450
            P_Fantravel = 130
            P_FRG = 300
            total_points = P_FT + P_OLKA + P_FRG + P_Fantravel
            current_points = 0
            
            # Progress Bar
            progress_bar = st.progress(0, text="Starter søgning...")
            status = st.status("Arbejder...", expanded=True)
            
            # --- 1. FootballTravel ---
            status.write("🤓 Data fra Footballtravel")
            time.sleep(1)  # Simpel pause for bedre UX
            try:
                df1 = Footballtravel.get_prices(selected)
                if not df1.empty: df1['Provider'] = "Footballtravel.dk"
                st.toast(f"Footballtravel: {len(df1)} tilbud fundet", icon="✅")
            except Exception as e:
                st.error(f"Fejl i Footballtravel: {e}")
                df1 = pd.DataFrame()
            current_points += P_FT
            progress_bar.progress(current_points / total_points, text="Footballtravel færdig...")
            
            # --- 2. Olka ---
            status.write("🌐 Data fra Olka")
            try:
                df2 = Olka.get_prices(selected)
                st.toast(f"Olka: {len(df2)} tilbud fundet", icon="✅" if not df2.empty else "⚠️")
            except Exception as e:
                st.error(f"Fejl i OLKA: {e}")
                df2 = pd.DataFrame()
            current_points += P_OLKA
            progress_bar.progress(current_points / total_points, text="Olka færdig...")

            # --- 3. Fantravel ---
            status.write("🤡 Data fra Fantravel")
            try:
                df3 = Fantravel.get_prices(selected)
                st.toast(f"Fantravel: {len(df3)} tilbud fundet", icon="✅" if not df3.empty else "⚠️")
            except Exception as e:
                st.error(f"Fejl i Fantravel: {e}")
                df3 = pd.DataFrame()
            current_points += P_Fantravel
            progress_bar.progress(current_points / total_points, text="Fantravel færdig...")

            # --- 4. Fodboldrejseguiden ---
            status.write("👽 Data fra resterende")
            try:
                df5 = Fodboldrejseguiden.get_prices(selected)
                st.toast(f"Fodboldrejseguiden: {len(df5)} tilbud fundet", icon="✅")
            except Exception as e:
                st.error(f"Fejl ved resterende: {e}")
                df5 = pd.DataFrame()
            current_points += P_FRG
            progress_bar.progress(1.0, text="Færdig!")

            # --- STOP TIMER ---
            end_time = time.time()
            elapsed = int(end_time - start_time)
            mins, secs = divmod(elapsed, 60)
            status.update(label=f"Færdig! (Tid: {mins}m {secs}s)", state="complete", expanded=False)
            st.success(f"✅ Søgning gennemført på {mins} minutter og {secs} sekunder.")

            # --- SAML DATA ---
            frames = [df1, df2, df3, df5]
            if all(df.empty for df in frames):
                st.warning("Ingen priser fundet.")
                st.stop()
            
            full_df = pd.concat(frames, ignore_index=True)

            # Rensning
            full_df['Provider'] = full_df['Provider'].fillna("Ukendt").astype(str)
            full_df = full_df[full_df['Provider'].str.strip() != ""]
            full_df['SortDate'] = pd.to_datetime(full_df['SortDate'], errors='coerce')
            full_df = full_df.dropna(subset=['SortDate'])

            # Filter: > 24 timer
            cutoff = datetime.now() + timedelta(hours=24)
            full_df = full_df[full_df['SortDate'] > cutoff]
            if full_df.empty:
                st.warning("Ingen relevante kampe fundet.")
                st.stop()

            # Sortering og ID-generering
            full_df = full_df.sort_values(by=['Club', 'SortDate'])
            full_df['club_change'] = full_df['Club'] != full_df['Club'].shift()
            full_df['date_diff'] = full_df['SortDate'].diff().dt.days.abs()
            full_df['big_gap'] = full_df['date_diff'] > 2 
            full_df['Match_Group_ID'] = (full_df['club_change'] | full_df['big_gap']).cumsum()


            # --- FORBERED DATA TIL EXCEL (TRANSFORMERING) ---
            
            # 1. Find alle unikke udbydere og sorter dem
            all_providers = sorted(full_df['Provider'].unique())
            if "Footballtravel.dk" in all_providers:
                all_providers.remove("Footballtravel.dk")
                all_providers.insert(0, "Footballtravel.dk")

            # 2. Gruppér data per kamp
            matches_grouped = full_df.groupby('Match_Group_ID').agg({
                'Club': 'first',
                'Match': lambda x: max(x, key=len),
                'SortDate': 'first'
            }).reset_index()

            match_data_list = []
            
            for _, match_row in matches_grouped.iterrows():
                group_id = match_row['Match_Group_ID']
                match_name = match_row['Match']
                date_str = match_row['SortDate'].strftime('%d/%m')
                display_name = f"{match_name} ({date_str})"
                
                prices_in_group = full_df[full_df['Match_Group_ID'] == group_id]
                
                provider_data = {}
                # Liste til at finde min/max for denne specifikke kamp
                valid_prices = [] 

                for _, p_row in prices_in_group.iterrows():
                    prov = p_row['Provider']
                    price_val = p_row['Price']
                    provider_data[prov] = {
                        'price': price_val,
                        'nights': p_row['Nights']
                    }
                    if price_val > 0:
                        valid_prices.append(price_val)
                
                # --- NYT: Beregn min og max pris for denne kamp ---
                min_price = min(valid_prices) if valid_prices else None
                max_price = max(valid_prices) if valid_prices else None

                match_data_list.append({
                    'display': display_name,
                    'data': provider_data,
                    'min_price': min_price, # Gem min pris
                    'max_price': max_price  # Gem max pris
                })

            # --- 7. EXCEL GENERERING ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                writer.book.create_sheet('Prices')
                ws = writer.book['Prices']
                
                # Definitioner af styles
                header_font = Font(bold=True)
                header_alignment = Alignment(textRotation=45, vertical='bottom', horizontal='center')
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                # --- NYT: Farve-definitioner ---
                # Light Green for billigst
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                # Light Red for dyrest
                red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                # --- A. SKRIV HEADERS ---
                col_idx = 2
                for match in match_data_list:
                    cell_match = ws.cell(row=1, column=col_idx, value=match['display'])
                    cell_match.font = header_font
                    cell_match.alignment = header_alignment
                    cell_match.border = thin_border
                    
                    cell_nights = ws.cell(row=1, column=col_idx+1, value="Nætter")
                    cell_nights.font = header_font
                    cell_nights.alignment = header_alignment
                    cell_nights.border = thin_border
                    
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                    ws.column_dimensions[get_column_letter(col_idx+1)].width = 8
                    
                    col_idx += 2

                # --- B. SKRIV RÆKKER (VIRKSOMHEDER) ---
                row_idx = 2
                
                for provider in all_providers:
                    cell_prov = ws.cell(row=row_idx, column=1, value=provider)
                    cell_prov.font = Font(bold=True)
                    cell_prov.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))
                    
                    col_idx = 2
                    for match in match_data_list:
                        p_data = match['data'].get(provider, {'price': 0, 'nights': 0})
                        price = p_data['price']
                        nights = p_data['nights']
                        
                        # Skriv Pris
                        cell_p = ws.cell(row=row_idx, column=col_idx, value=price if price > 0 else "")
                        cell_p.border = thin_border
                        
                        # --- NYT: Logik for farvning af pris ---
                        if price > 0:
                            # Er det den billigste? (Prioriteret)
                            if price == match['min_price']:
                                cell_p.fill = green_fill
                            # Er det den dyreste? (Og sørg for vi ikke farver rød hvis der kun er én pris, som også er min)
                            elif price == match['max_price']:
                                cell_p.fill = red_fill
                        
                        # Skriv Nætter
                        cell_n = ws.cell(row=row_idx, column=col_idx+1, value=nights if nights > 0 else "")
                        cell_n.border = thin_border
                        
                        col_idx += 2
                    
                    row_idx += 1

                ws.column_dimensions['A'].width = 25
                ws.freeze_panes = "B2"

            # Download Knap og Preview (uændret)
            timestamp = datetime.now().strftime("%H-%M")
            st.download_button(
                "📥 Download Excel", 
                output.getvalue(), 
                f"prices_matrix_{timestamp}.xlsx", 
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Vis preview i Streamlit (Vi laver en simpel dataframe til visning da Streamlit ikke viser rotationer)
            preview_df = pd.DataFrame(index=all_providers)
            for m in match_data_list:
                col_name = m['display']
                # Byg en kolonne med priser for preview
                prices = []
                for p in all_providers:
                    val = m['data'].get(p, {}).get('price', 0)
                    prices.append(val if val > 0 else 0)
                preview_df[col_name] = prices
            
            st.write("Preview af data:")
            st.dataframe(preview_df, use_container_width=True)

if __name__ == "__main__":
    main()